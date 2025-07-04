"""
Utility commands cog for the Twi Bot Shard.

This module provides a variety of utility commands for server management, user information,
role management, and other miscellaneous functionality. It includes commands for user info,
server info, role management, quotes, dice rolling, and more.
"""

# Import standard library modules first
import logging
import os
import re
import time
import random
from datetime import datetime
from itertools import groupby
from os.path import exists
from typing import List

# Import third-party modules in a specific order
# Discord-related imports
import discord
from discord import app_commands
from discord.ext import commands

# Other third-party imports
from openpyxl import load_workbook, Workbook

# Import AO3 last to avoid potential import deadlocks
# Adding a small delay before AO3 import to prevent deadlocks
import time

time.sleep(0.1)  # 100ms delay to avoid potential race conditions
import AO3

import config
from utils.permissions import (
    admin_or_me_check,
    admin_or_me_check_wrapper,
    app_admin_or_me_check,
)

LOGIN_AO3_SUCCESSFUL = False
try:
    session = AO3.Session(str(config.ao3_username), str(config.ao3_password))
    LOGIN_AO3_SUCCESSFUL = True
except Exception as e:
    logging.error(f"AO3 login error: {e}")


async def user_info_function(interaction: discord.Interaction, member: discord.Member):
    """
    Create and send an embed with detailed information about a Discord user.

    This function creates an embed containing information about a user, including
    their account creation date, server join date, ID, color, and roles. It's used
    by both the /info user command and the User Info context menu.

    Args:
        interaction: The interaction that triggered the command
        member: The member to get information about, defaults to the command user if None
    """
    if member is None:
        member = interaction.user
    embed = discord.Embed(title=member.display_name, color=discord.Color(0x3CD63D))
    embed.set_thumbnail(url=member.display_avatar.url)
    embed.add_field(
        name="Account created at",
        value=member.created_at.strftime("%d-%m-%Y @ %H:%M:%S"),
    )
    embed.add_field(
        name="Joined server", value=member.joined_at.strftime("%d-%m-%Y @ %H:%M:%S")
    )
    embed.add_field(name="Id", value=member.id)
    embed.add_field(name="Color", value=member.color)
    roles = ""
    for role in reversed(member.roles):
        if role.is_default():
            pass
        else:
            roles += f"{role.mention}\n"
    if roles != "":
        embed.add_field(name="Roles", value=roles, inline=False)
    await interaction.response.send_message(embed=embed)


class OtherCogs(commands.Cog, name="Other"):
    """
    Cog providing various utility commands for server management and user interaction.

    This cog includes commands for user information, server information, role management,
    quotes, dice rolling, and other utility functions. It also provides context menu
    commands for pinning messages and viewing user information.

    Attributes:
        bot: The bot instance
        quote_cache: Cache of quotes for autocomplete
        category_cache: Cache of role categories for autocomplete
        pin_cache: Cache of channels where pinning is allowed
    """

    def __init__(self, bot: commands.Bot) -> None:
        """
        Initialize the OtherCogs cog.

        Args:
            bot: The bot instance to which this cog is attached
        """
        self.bot: commands.Bot = bot
        self.logger = logging.getLogger('cogs.other')
        self.quote_cache = None
        self.category_cache = None
        self.pin_cache = None

        self.pin = app_commands.ContextMenu(
            name="Pin",
            callback=self.pin,
        )
        self.bot.tree.add_command(self.pin)

        self.info_user_context = app_commands.ContextMenu(
            name="User info",
            callback=self.info_user_context,
        )
        self.bot.tree.add_command(self.info_user_context)

    async def cog_load(self) -> None:
        """
        Load initial data when the cog is added to the bot.

        This method is called automatically when the cog is loaded.
        It populates the quote cache, category cache, and pin cache
        for use in commands and autocomplete.
        """
        self.quote_cache = await self.bot.db.fetch(
            "SELECT quote, row_number FROM (SELECT quote, ROW_NUMBER () OVER () FROM quotes) x"
        )
        self.category_cache = await self.bot.db.fetch(
            "SELECT DISTINCT (category) FROM roles WHERE category IS NOT NULL"
        )
        self.pin_cache = await self.bot.db.fetch(
            "SELECT id FROM channels where allow_pins = TRUE"
        )

    @app_commands.command(
        name="ping",
        description="Gives the latency of the bot",
    )
    async def ping(self, interaction: discord.Interaction) -> None:
        """
        Display the bot's current latency to Discord.

        This command calculates and displays the bot's WebSocket latency to Discord
        in milliseconds, which can be useful for diagnosing connection issues.

        Args:
            interaction: The interaction that triggered the command
        """
        await interaction.response.send_message(f"{round(self.bot.latency * 1000)} ms")

    @app_commands.command(
        name="avatar", description="Posts the full version of a avatar"
    )
    async def av(
        self, interaction: discord.Interaction, member: discord.Member = None
    ) -> None:
        """
        Display the full-size avatar of a user.

        This command creates an embed containing the full-size version of a user's
        avatar, which can be useful for viewing avatars in higher resolution.

        Args:
            interaction: The interaction that triggered the command
            member: The member whose avatar to display, defaults to the command user if None
        """
        if member is None:
            member = interaction.user
        embed = discord.Embed(title="Avatar", color=discord.Color(0x3CD63D))
        embed.set_image(url=member.display_avatar.url)
        await interaction.response.send_message(embed=embed)

    info = app_commands.Group(name="info", description="Information commands")

    @info.command(
        name="user",
        description="Gives the account information of a user.",
    )
    async def info_user(
        self, interaction: discord.Interaction, member: discord.Member = None
    ) -> None:
        """
        Display detailed information about a Discord user.

        This command shows information about a user, including their account creation date,
        server join date, ID, color, and roles.

        Args:
            interaction: The interaction that triggered the command
            member: The member to get information about, defaults to the command user if None
        """
        await user_info_function(interaction, member)

    async def info_user_context(
        self, interaction: discord.Interaction, member: discord.Member
    ) -> None:
        """
        Context menu command to display detailed information about a Discord user.

        This context menu command shows the same information as the /info user command,
        but can be accessed by right-clicking on a user.

        Args:
            interaction: The interaction that triggered the command
            member: The member to get information about
        """
        await user_info_function(interaction, member)

    @info.command(
        name="server",
        description="Gives the server information of the server the command was used in.",
    )
    async def info_server(self, interaction: discord.Interaction) -> None:
        """
        Display detailed information about the current Discord server.

        This command shows comprehensive information about the server, including
        its creation date, owner, member count, role count, emoji counts, and more.

        Args:
            interaction: The interaction that triggered the command
        """
        embed = discord.Embed(
            title=interaction.guild.name, color=discord.Color(0x3CD63D)
        )
        embed.set_thumbnail(url=interaction.guild.icon)
        embed.add_field(name="Id", value=interaction.guild.id, inline=False)
        embed.add_field(name="Owner", value=interaction.guild.owner.mention)
        embed.add_field(
            name="Created At",
            value=interaction.guild.created_at.strftime("%d-%m-%Y @ %H:%M:%S"),
            inline=False,
        )
        embed.add_field(name="Banner", value=interaction.guild.banner, inline=False)
        embed.add_field(
            name="Description", value=interaction.guild.description, inline=False
        )
        embed.add_field(name="Member count", value=interaction.guild.member_count)
        embed.add_field(
            name="Number of roles", value=f"{len(interaction.guild.roles)}/250"
        )
        normal, animated = 0, 0
        for emoji in interaction.guild.emojis:
            if emoji.animated:
                animated += 1
            else:
                normal += 1
        embed.add_field(
            name="Number of emojis", value=f"{normal}/{interaction.guild.emoji_limit}"
        )
        embed.add_field(
            name="Number of Animated emojis",
            value=f"{animated}/{interaction.guild.emoji_limit}",
        )
        embed.add_field(
            name="Number of stickers",
            value=f"{len(interaction.guild.stickers)}/{interaction.guild.sticker_limit}",
        )
        embed.add_field(
            name="Number of active threads",
            value=f"{len(await interaction.guild.active_threads())}",
        )
        embed.add_field(
            name="Number of Text Channels",
            value=f"{len(interaction.guild.text_channels)}",
        )
        embed.add_field(
            name="Number of Voice channels",
            value=f"{len(interaction.guild.voice_channels)}",
        )
        if interaction.guild.vanity_url is not None:
            embed.add_field(name="Invite link", value=interaction.guild.vanity_url)
        await interaction.response.send_message(embed=embed)

    @info.command(
        name="role", description="Gives the role information of the role given."
    )
    async def info_role(
        self, interaction: discord.Interaction, role: discord.Role
    ) -> None:
        """
        Display detailed information about a Discord role.

        This command shows information about a role, including its color,
        creation date, whether it's hoisted (displayed separately), ID, and
        the number of members who have the role.

        Args:
            interaction: The interaction that triggered the command
            role: The role to get information about
        """
        embed = discord.Embed(title=role.name, color=(discord.Color(role.color.value)))
        embed.add_field(name="Color: ", value=hex(role.color.value), inline=False)
        embed.add_field(
            name="Created at",
            value=role.created_at.strftime("%d-%m-%Y @ %H:%M:%S"),
            inline=False,
        )
        embed.add_field(name="Hoisted", value=role.hoist, inline=False)
        embed.add_field(name="Id", value=role.id, inline=False)
        embed.add_field(name="Member count", value=len(role.members), inline=False)
        await interaction.response.send_message(embed=embed)

    @app_commands.command(
        name="say",
        description="Makes Cognita repeat whatever was said",
    )
    @commands.is_owner()
    async def say(self, interaction: discord.Interaction, say: str) -> None:
        """
        Make the bot repeat a message in the current channel.

        This owner-only command makes the bot send a message with the specified content
        in the current channel. The command response is ephemeral (only visible to the
        command user) to avoid cluttering the channel.

        Args:
            interaction: The interaction that triggered the command
            say: The message content to repeat
        """
        await interaction.response.send_message("Sent message", ephemeral=True)
        await interaction.channel.send(say)

    @app_commands.command(
        name="saychannel",
        description="Makes Cognita repeat whatever was said in a specific channel",
    )
    @commands.is_owner()
    async def say_channel(
        self, interaction: discord.Interaction, channel: discord.TextChannel, say: str
    ) -> None:
        """
        Make the bot repeat a message in a specified channel.

        This owner-only command makes the bot send a message with the specified content
        in the specified channel. The command response is ephemeral (only visible to the
        command user) to avoid cluttering the channel.

        Args:
            interaction: The interaction that triggered the command
            channel: The channel to send the message in
            say: The message content to repeat
        """
        await interaction.response.send_message("Sent message", ephemeral=True)
        await channel.send(say)

    @commands.Cog.listener()
    async def on_member_update(
        self, before: discord.Member, after: discord.Member
    ) -> None:
        """
        Event handler for when a member's roles are updated.

        This listener detects when a member gains a special role and sends a themed
        announcement message to the inn general channel. Each special role has a
        unique announcement message themed around The Wandering Inn.

        Args:
            before: The member's state before the update
            after: The member's state after the update
        """
        # Acid jars, Acid Flies, Frying Pans, Enchanted Soup, Barefoot Clients.
        # Green, Purple, Orange, Blue, Red
        list_of_ids = [
            config.special_role_ids["acid_jars"],
            config.special_role_ids["acid_flies"],
            config.special_role_ids["frying_pans"],
            config.special_role_ids["enchanted_soup"],
            config.special_role_ids["barefoot_clients"],
        ]
        gained = set(after.roles) - set(before.roles)
        if gained:
            gained = gained.pop()
            if gained.id in list_of_ids:
                channel = self.bot.get_channel(config.inn_general_channel_id)
                # Acid jar
                if gained.id == config.special_role_ids["acid_jars"]:
                    embed = discord.Embed(
                        title=f"Hey be careful over there!",
                        description=f"Those {gained.mention} will melt your hands off {after.mention}!",
                    )
                # Acid Flies
                elif gained.id == config.special_role_ids["acid_flies"]:
                    embed = discord.Embed(
                        title=f"Make some room at the tables!",
                        description=f"{after.mention} just ordered a bowl of {gained.mention}!",
                    )
                # Frying Pans
                elif gained.id == config.special_role_ids["frying_pans"]:
                    embed = discord.Embed(
                        title=f"Someone ordered a frying pan!",
                        description=f"Hope {after.mention} can dodge!",
                    )
                # Enchanted Soup
                elif gained.id == config.special_role_ids["enchanted_soup"]:
                    embed = discord.Embed(
                        title=f"Hey get down from there Mrsha!",
                        description=f"Looks like {after.mention} will have to order a new serving of {gained.mention} because Mrsha just ate theirs!",
                    )
                # Barefoot Clients
                elif gained.id == config.special_role_ids["barefoot_clients"]:
                    embed = discord.Embed(
                        title=f"Make way!",
                        description=f"{gained.mention} {after.mention} coming through!",
                    )
                else:
                    embed = discord.Embed(
                        title=f"Make some room in the inn!",
                        description=f"{after.mention} just joined the ranks of {gained.mention}!",
                    )
                embed.set_thumbnail(url=after.display_avatar.url)
                await channel.send(embed=embed, content=f"{after.mention}")

    quote = app_commands.Group(name="quote", description="Quote commands")

    @quote.command(name="add", description="Adds a quote to the list of quotes")
    async def quote_add(self, interaction: discord.Interaction, quote: str):
        """
        Add a new quote to the database.

        This command adds a new quote to the database, recording the author's name,
        ID, and the current timestamp. It also updates the quote cache for autocomplete.

        Args:
            interaction: The interaction that triggered the command
            quote: The quote text to add
        """
        await self.bot.db.execute(
            "INSERT INTO quotes(quote, author, author_id, time, tokens) VALUES ($1,$2,$3,now(),to_tsvector($4))",
            quote,
            interaction.user.display_name,
            interaction.user.id,
            quote,
        )
        row_number = await self.bot.db.fetchrow("SELECT COUNT(*) FROM quotes")
        await interaction.response.send_message(
            f"Added quote `{quote}` at index {row_number['count']}"
        )
        self.quote_cache = await self.bot.db.fetch(
            "SELECT quote, row_number FROM (SELECT quote, ROW_NUMBER () OVER () FROM quotes) x"
        )

    @quote.command(name="find", description="Searches for a quote")
    async def quote_find(self, interaction: discord.Interaction, search: str):
        """
        Search for quotes containing specific words.

        This command searches the quotes database for quotes containing the specified
        search terms. It uses PostgreSQL's full-text search capabilities for efficient
        searching. If multiple results are found, it displays the first result and
        indicates that there are more results available.

        Args:
            interaction: The interaction that triggered the command
            search: The search terms to look for in quotes
        """
        formatted_search = search.replace(" ", " & ")
        results = await self.bot.db.fetch(
            "SELECT quote, x.row_number FROM (SELECT tokens, quote, ROW_NUMBER() OVER () as row_number FROM quotes) x WHERE x.tokens @@ to_tsquery($1);",
            formatted_search,
        )

        if len(results) > 1:
            index_res = "["
            iterres = iter(results)
            next(iterres)
            for result in iterres:
                index_res = f"{index_res}{str(result['row_number'])}, "
            index_res = index_res[:-2]
            index_res = f"{index_res}]"
            await interaction.response.send_message(
                f"Quote {results[0]['row_number']}: {results[0]['quote']}\nThere is also results at {index_res}"
            )
        elif len(results) == 1:
            await interaction.response.send_message(
                f"Quote {results[0]['row_number']}: {results[0]['quote']}"
            )
        elif len(results) < 1:
            await interaction.response.send_message("I found no results")
        else:
            await interaction.response.send_message("How the fuck?")

    @quote.command(name="delete", description="Delete a quote")
    async def quote_delete(self, interaction: discord.Interaction, delete: int):
        """
        Delete a quote from the database by its index.

        This command removes a quote from the database based on its row number.
        It also updates the quote cache for autocomplete after deletion.

        Args:
            interaction: The interaction that triggered the command
            delete: The row number of the quote to delete
        """
        u_quote = await self.bot.db.fetchrow(
            "SELECT quote, row_number FROM (SELECT quote, ROW_NUMBER () OVER () FROM quotes) x WHERE ROW_NUMBER = $1",
            delete,
        )
        if u_quote:
            await self.bot.db.execute(
                "DELETE FROM quotes WHERE serial_id in (SELECT serial_id FROM QUOTES ORDER BY TIME LIMIT 1 OFFSET $1)",
                delete - 1,
            )
            await interaction.response.send_message(
                f"Deleted quote `{u_quote['quote']}` from position {u_quote['row_number']}"
            )
            self.quote_cache = await self.bot.db.fetch(
                "SELECT quote, row_number FROM (SELECT quote, ROW_NUMBER () OVER () FROM quotes) x"
            )
        else:
            await interaction.response.send_message(
                "Im sorry. I could not find a quote on that index"
            )

    @quote_delete.autocomplete("delete")
    async def quote_delete_autocomplete(
        self, interaction: discord.Interaction, current: int
    ) -> List[app_commands.Choice[int]]:
        """
        Provide autocomplete suggestions for the quote delete command.

        This method filters the cached quotes based on the user's current input
        and returns matching options for the autocomplete dropdown.

        Args:
            interaction: The interaction that triggered the autocomplete
            current: The current number input by the user

        Returns:
            A list of up to 25 matching quote choices with their row numbers
        """
        ln = []
        for x in self.quote_cache:
            ln.append({"quote": x["quote"], "row_number": x["row_number"]})
        return [
            app_commands.Choice(
                name=f"{quote['row_number']}: {quote['quote']}"[0:100],
                value=quote["row_number"],
            )
            for quote in ln
            if str(current) in str(quote["row_number"]) or current == ""
        ][0:25]

    @quote.command(
        name="get",
        description="Posts a quote a random quote or a quote with the given index",
    )
    async def quote_get(self, interaction, index: int = None):
        """
        Retrieve and display a quote from the database.

        This command retrieves either a random quote or a specific quote by its
        row number. If no index is provided, a random quote is selected.

        Args:
            interaction: The interaction that triggered the command
            index: Optional row number of the quote to retrieve, random if None
        """
        if index is None:
            u_quote = await self.bot.db.fetchrow(
                "SELECT quote, row_number FROM (SELECT quote, ROW_NUMBER () OVER () FROM quotes) x  ORDER BY random() LIMIT 1"
            )
        else:
            u_quote = await self.bot.db.fetchrow(
                "SELECT quote, row_number FROM (SELECT quote, ROW_NUMBER () OVER () FROM quotes) x WHERE ROW_NUMBER = $1",
                index,
            )
        if u_quote:
            await interaction.response.send_message(
                f"Quote {u_quote['row_number']}: `{u_quote['quote']}`"
            )
        else:
            await interaction.response.send_message(
                "Im sorry, i could not find a quote with that index value."
            )

    @quote_get.autocomplete("index")
    async def quote_get_autocomplete(
        self,
        interaction,
        current: int,
    ) -> List[app_commands.Choice[int]]:
        """
        Provide autocomplete suggestions for the quote get command.

        This method filters the cached quotes based on the user's current input
        and returns matching options for the autocomplete dropdown.

        Args:
            interaction: The interaction that triggered the autocomplete
            current: The current number input by the user

        Returns:
            A list of up to 25 matching quote choices with their row numbers
        """
        ln = []
        for x in self.quote_cache:
            ln.append({"quote": x["quote"], "row_number": x["row_number"]})
        return [
            app_commands.Choice(
                name=f"{quote['row_number']}: {quote['quote']}"[0:100],
                value=quote["row_number"],
            )
            for quote in ln
            if str(current) in str(quote["row_number"]) or current == ""
        ][0:25]

    @quote.command(name="who", description="Posts who added a quote")
    async def quote_who(self, interaction, index: int):
        """
        Display information about who added a specific quote.

        This command retrieves metadata about a quote, including the username
        and ID of the user who added it, and when it was added.

        Args:
            interaction: The interaction that triggered the command
            index: The row number of the quote to get information about
        """
        u_quote = await self.bot.db.fetchrow(
            "SELECT author, author_id, time, row_number FROM (SELECT author, author_id, time, ROW_NUMBER () OVER () FROM quotes) x WHERE ROW_NUMBER = $1",
            index,
        )
        if u_quote:
            await interaction.response.send_message(
                f"Quote {u_quote['row_number']} was added by: {u_quote['author']} ({u_quote['author_id']}) at {u_quote['time']}"
            )
        else:
            await interaction.response.send_message(
                "Im sorry, i could not find a quote with that index value."
            )

    @quote_who.autocomplete("index")
    async def quote_who_autocomplete(
        self,
        interaction,
        current: int,
    ) -> List[app_commands.Choice[int]]:
        """
        Provide autocomplete suggestions for the quote who command.

        This method filters the cached quotes based on the user's current input
        and returns matching options for the autocomplete dropdown.

        Args:
            interaction: The interaction that triggered the autocomplete
            current: The current number input by the user

        Returns:
            A list of up to 25 matching quote choices with their row numbers
        """
        ln = []
        for x in self.quote_cache:
            ln.append({"quote": x["quote"], "row_number": x["row_number"]})
        return [
            app_commands.Choice(
                name=f"{quote['row_number']}: {quote['quote']}"[0:100],
                value=quote["row_number"],
            )
            for quote in ln
            if str(current) in str(quote["row_number"]) or current == ""
        ][0:25]

    @app_commands.command(
        name="roles",
        description="Posts all the roles in the server you can assign yourself",
    )
    async def role_list(self, interaction: discord.Interaction):
        """
        Display a list of all self-assignable roles in the server.

        This command creates an embed containing all roles that users can assign
        to themselves using the /role command. Roles are grouped by category and
        displayed as mentions for easy assignment.

        Args:
            interaction: The interaction that triggered the command
        """
        user_roles = [role.id for role in interaction.user.roles]
        roles = await self.bot.db.fetch(
            "SELECT id, name, required_roles, weight, category "
            "FROM roles "
            "WHERE (required_roles && $2::bigint[] OR required_roles is NULL)"
            "AND guild_id = $1 "
            "AND self_assignable = TRUE "
            "order by weight, name desc",
            interaction.guild.id,
            user_roles,
        )

        if not roles:
            await interaction.response.send_message(
                "Doesn't look like any roles has been setup to be self assignable on this server."
                "The moderators can do that by using: "
                "`/addrole [Role]`"
            )
            return

        embed = discord.Embed(
            title="List of all the roles in the server",
            description="Request the role by doing /role @Role",
            color=0x00FCFF,
        )
        embed.set_thumbnail(url=interaction.guild.icon)
        roles.sort(key=lambda k: (k["category"], k["weight"]))
        for key, group in groupby(roles, key=lambda k: k["category"]):
            role_mentions = ""
            x = 1
            for row in group:
                role = interaction.guild.get_role(row["id"])
                if role:
                    temp_str = f"{role.mention}\n"
                    if len(temp_str + role_mentions) > 1024:
                        embed.add_field(
                            name=f"**{key.capitalize()}**",
                            value=role_mentions.strip(),
                            inline=False,
                        )
                        role_mentions = ""
                        x += 1
                        key = f"{key} {x}"
                    role_mentions = role_mentions + temp_str
            embed.add_field(
                name=f"**{key.capitalize()}**",
                value=role_mentions.strip(),
                inline=False,
            )
        await interaction.response.send_message(embed=embed)

    admin_role = app_commands.Group(
        name="admin_role", description="Admin role commands"
    )

    @admin_role.command(name="weight", description="Changes the weight of a role")
    @app_commands.check(app_admin_or_me_check)
    async def update_role_weight(
        self, interaction: discord.Interaction, role: discord.role.Role, new_weight: int
    ):
        """
        Change the weight of a role in the role list.

        This admin-only command updates the weight of a role, which affects its
        position in the role list when displayed with the /roles command.
        Roles with lower weights appear higher in the list within their category.

        Args:
            interaction: The interaction that triggered the command
            role: The role to update
            new_weight: The new weight value to assign to the role
        """
        await self.bot.db.execute(
            "UPDATE roles set weight = $1 WHERE id = $2 AND guild_id = $3",
            new_weight,
            role.id,
            interaction.guild.id,
        )
        await interaction.response.send_message(
            f"Changed the weight of {role} to {new_weight}"
        )

    @admin_role.command(name="add", description="Adds a role to the self assign list")
    @app_commands.check(app_admin_or_me_check)
    async def role_add(
        self,
        interaction: discord.Interaction,
        role: discord.role.Role,
        category: str = "Uncategorized",
        auto_replace: bool = False,
        required_roles: str = None,
    ):
        """
        Add a role to the self-assignable roles list.

        This admin-only command makes a role self-assignable by users with the /role command.
        It allows specifying a category for organization, whether the role should automatically
        replace other roles in the same category, and any roles required to access this role.

        Args:
            interaction: The interaction that triggered the command
            role: The role to make self-assignable
            category: The category to place the role in (default: 'Uncategorized')
            auto_replace: Whether this role should replace other roles in the same category
            required_roles: Space-separated list of role mentions or IDs required to access this role
        """
        try:
            if required_roles is not None:
                list_of_roles = []
                required_roles = required_roles.split(" ")
                for user_role in required_roles:
                    matched_id = re.search(r"\d+", user_role)
                    if matched_id:
                        temp = discord.utils.get(
                            interaction.guild.roles, id=matched_id.group()
                        )
                        if temp:
                            list_of_roles.append(temp.id)
                await self.bot.db.execute(
                    "UPDATE roles SET self_assignable = TRUE, required_roles = $1, alias = $2, id = $3, guild_id = $4, category=$5, auto_replace = $6 "
                    "WHERE id = $3 AND guild_id = $4",
                    list_of_roles,
                    role.id,
                    interaction.guild.id,
                    category.lower(),
                    auto_replace,
                )
            else:
                await self.bot.db.execute(
                    "UPDATE roles SET self_assignable = TRUE, required_roles = NULL, alias = $1, id = $2, guild_id = $3, category=$4, auto_replace = $5 "
                    "WHERE id = $2 AND guild_id = $3",
                    role.id,
                    interaction.guild.id,
                    category.lower(),
                    auto_replace,
                )
            await interaction.response.send_message(
                f"Added {role} to the self assign list"
            )
        except Exception as e:
            logging.exception(e)
            await interaction.response.send_message(f"Error: {e}")

    @role_add.autocomplete("category")
    async def role_add_autocomplete(
        self,
        interaction: discord.Interaction,
        current: str,
    ) -> List[app_commands.Choice[str]]:
        """
        Provide autocomplete suggestions for role categories.

        This method filters the cached role categories based on the user's current input
        and returns matching options for the autocomplete dropdown.

        Args:
            interaction: The interaction that triggered the autocomplete
            current: The current text input by the user

        Returns:
            A list of up to 25 matching category name choices
        """
        return [
            app_commands.Choice(name=category, value=category)
            for category in self.category_cache
            if current.lower() in category.lower() or current == ""
        ][0:25]

    @admin_role.command(
        name="remove", description="removes a role from the self assign list"
    )
    @app_commands.check(app_admin_or_me_check)
    async def role_remove(self, interaction: discord.Interaction, role: discord.Role):
        """
        Remove a role from the self-assignable roles list.

        This admin-only command removes a role from the self-assignable roles list,
        preventing users from assigning it to themselves with the /role command.
        It resets all role settings to their defaults.

        Args:
            interaction: The interaction that triggered the command
            role: The role to remove from the self-assignable list
        """
        await self.bot.db.execute(
            "UPDATE roles SET self_assignable = FALSE, weight = 0, alias = NULL, category = NULL, required_role = NULL, auto_replace = FALSE "
            "where id = $1 "
            "AND guild_id = $2",
            role,
            interaction.guild.id,
        )
        await interaction.response.send_message(
            f"Removed {role} from the self assign list"
        )

    @app_commands.command(
        name="role", description="Adds or removes a role from yourself"
    )
    async def role(self, interaction: discord.Interaction, role: discord.Role):
        """
        Add or remove a self-assignable role from yourself.

        This command toggles a self-assignable role on the user who runs it.
        If the user already has the role, it will be removed. If they don't have it,
        it will be added. The command checks if the role is self-assignable and if
        the user has any required roles needed to access it.

        If the role has auto_replace enabled, any other roles in the same category
        will be removed when this role is added.

        Args:
            interaction: The interaction that triggered the command
            role: The role to add or remove
        """
        s_role = await self.bot.db.fetchrow(
            "SELECT * FROM roles WHERE id = $1", role.id
        )
        if s_role["self_assignable"]:
            b_role = list()
            for a_role in interaction.user.roles:
                b_role.append(a_role.id)
            if (
                s_role["required_roles"] is None
                or [i for i in s_role["required_roles"] if i in b_role] != []
            ):
                if role in interaction.user.roles:
                    await interaction.user.remove_roles(role)
                    await interaction.response.send_message(f"I removed {role}")
                else:
                    if await self.bot.db.fetchval(
                        "SELECT auto_replace FROM roles WHERE id = $1", role.id
                    ):
                        list_r = list()
                        for r in interaction.user.roles:
                            list_r.append(r.id)
                        r_loop = self.bot.db.fetch(
                            "SELECT id FROM roles WHERE id = ANY($1::bigint[]) "
                            "AND category = "
                            "(SELECT category FROM roles where id = $2)",
                            list_r,
                            role.id,
                        )
                        for r in await r_loop:
                            await interaction.user.remove_roles(
                                interaction.guild.get_role(r["id"])
                            )
                    await interaction.user.add_roles(role)
                    await interaction.response.send_message(f"I added {role}")
            else:
                await interaction.response.send_message(
                    "You do not have the required role for that."
                )
        else:
            await interaction.response.send_message(
                "The requested role is not self assignable"
            )

    @app_commands.command(name="roll", description="Rolls a dice")
    async def roll(
        self,
        interaction: discord.Interaction,
        dice: int = 20,
        amount: int = 1,
        modifier: int = 0,
    ):
        """
        Roll dice and display the results.

        This command simulates rolling dice with a specified number of sides,
        amount of dice, and an optional modifier added to the total.
        By default, it rolls a single 20-sided die (d20) with no modifier.

        Args:
            interaction: The interaction that triggered the command
            dice: The number of sides on each die (default: 20)
            amount: The number of dice to roll (default: 1, max: 100)
            modifier: A number to add to the total roll (default: 0)
        """
        if amount > 100:
            await interaction.response.send_message("I can't roll more than 100 dice")
        else:
            rolls = list()
            for x in range(amount):
                rolls.append(random.randint(1, dice))
            await interaction.response.send_message(
                f"Rolled {amount}d{dice} + {modifier} = {sum(rolls) + modifier} ({rolls})"
            )

    @app_commands.command(name="gallery_stats", description="Posts the gallery stats")
    async def gallery_stats(self, interaction: discord.Interaction):
        """
        Generate an Excel spreadsheet with gallery statistics.

        This command creates an Excel spreadsheet containing information about
        gallery submissions, including titles, links, creators, and posting dates.
        It processes all messages in the gallery channel that contain embeds.

        Args:
            interaction: The interaction that triggered the command
        """
        gallery = interaction.guild.get_channel_or_thread(964519175320125490)
        if exists("gallery.xlsx"):
            os.remove("gallery.xlsx")
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Entry #"
        ws["B1"] = "Title (if given)"
        ws["C1"] = "Link"
        ws["D1"] = "Fanwork Message link"
        ws["E1"] = (
            "Type of Submission (Classify based on posting location-- gallery or nsfw-gallery)"
        )
        ws["F1"] = "Creator (Use their Discord username, not their handle)"
        ws["G1"] = "Posted Date"
        ws["H1"] = "Inktober Prompt"
        ws["I1"] = "Quest( if applicable)"
        ws["J1"] = "Name"
        ws["K1"] = "Notes"
        ws["L1"] = "Image"
        first = datetime.strptime("2022-09-14", "%Y-%m-%d")
        row = 2
        async for message in gallery.history(
            limit=None, after=first, oldest_first=True
        ):
            if len(message.embeds) != 0 and message.author.id == 631257798465945650:
                logging.info(f"Message: {message.embeds[0].description}")
                # ws.cell(row=row, column=1).value = entry
                ws.cell(row=row, column=2).value = message.embeds[0].title
                ws.cell(row=row, column=3).value = message.embeds[0].image.url
                ws.cell(row=row, column=4).value = re.findall(
                    r"(https://discord\.com/channels/\d*/\d*/\d*)|$",
                    message.embeds[0].description,
                )[0]
                # ws.cell(row=row, column=5).value = 'gallery'
                ws.cell(row=row, column=6).value = interaction.guild.get_member(
                    int(re.findall(r"<@(\d*)>|$", message.embeds[0].description)[0])
                ).name
                ws.cell(row=row, column=7).value = message.created_at.strftime(
                    "%d-%m-%Y @ %H:%M:%S"
                )
                # ws.cell(row=row, column=8).value = message.embeds[0].description
                # ws.cell(row=row, column=9).value = message.embeds[0].fields[0].value
                # ws.cell(row=row, column=10).value = message.embeds[0].fields[1].value
                # ws.cell(row=row, column=11).value = message.embeds[0].fields[2].value
                # ws.cell(row=row, column=12).value = message.embeds[0].image.url
                wb.save("gallery.xlsx")
                row += 1

    @app_commands.command(name="ao3", description="Posts information about a ao3 work")
    async def ao3(self, interaction: discord.Interaction, ao3_url: str) -> None:
        """
        Display detailed information about an Archive of Our Own (AO3) work.

        This command retrieves and displays comprehensive information about a fanfiction
        work from AO3, including the title, author, summary, ratings, tags, statistics,
        and other metadata. It requires a valid AO3 URL to function.

        Args:
            interaction: The interaction that triggered the command
            ao3_url: The URL of the AO3 work to retrieve information about
        """
        if re.search(r"https?://archiveofourown.org/works/\d+", ao3_url):
            session.refresh_auth_token()
            ao3_id = AO3.utils.workid_from_url(ao3_url)
            try:
                work = AO3.Work(ao3_id)
            except AO3.utils.InvalidIdError:
                await interaction.response.send_message(
                    "I could not find that work on AO3", ephemeral=True
                )
                return
            work.set_session(session)
            embed = discord.Embed(
                title=work.title,
                description=work.summary,
                color=discord.Color(0x3CD63D),
                url=work.url,
            )
            new_line = "\n"
            try:
                authors = ""
                for author in work.authors:
                    author_name = re.search(
                        r"https?://archiveofourown.org/users/(\w+)", author.url
                    ).group(1)
                    authors += f"[{author_name}]({author.url})\n"
                embed.add_field(name="Author", value=authors[0:2000])
                embed.add_field(name="Rating", value=work.rating)
                embed.add_field(name="Category", value=f"{','.join(work.categories)}")
                embed.add_field(
                    name="Fandoms", value=f"{new_line.join(work.fandoms)[0:2000]}"
                )
                embed.add_field(
                    name="Relationships",
                    value=f"{new_line.join(work.relationships)[0:2000]}",
                )
                embed.add_field(
                    name="Characters", value=f"{new_line.join(work.characters)[0:2000]}"
                )
                embed.add_field(
                    name="Warnings", value=f"{new_line.join(work.warnings)}"
                )
                embed.add_field(name="Language", value=work.language)
                embed.add_field(name="Words", value=f"{int(work.words):,}")
                embed.add_field(
                    name="Chapters",
                    value=f"{work.nchapters}/{work.expected_chapters if work.expected_chapters is not None else '?'}",
                )
                embed.add_field(name="Comments", value=work.comments)
                embed.add_field(name="Kudos", value=work.kudos)
                embed.add_field(name="Bookmarks", value=work.bookmarks)
                embed.add_field(name="Hits", value=work.hits)
                embed.add_field(
                    name="Published", value=work.date_published.strftime("%Y-%m-%d")
                )
                embed.add_field(
                    name="Updated", value=work.date_updated.strftime("%Y-%m-%d")
                )
                embed.add_field(name="Status", value=work.status)
                embed.add_field(name="URL", value=work.url)
                await interaction.response.send_message(embed=embed)
            except AttributeError as e:
                logging.warning(f"AO3: {e}")
                await interaction.response.send_message(
                    "I could not find that work on AO3", ephemeral=True
                )
        else:
            await interaction.response.send_message(
                "That doesn't look like a link to ao3", ephemeral=True
            )

    # context menu command to pin a message
    async def pin(
        self, interaction: discord.Interaction, message: discord.Message
    ) -> None:
        """
        Context menu command to pin a message in allowed channels.

        This command allows users to pin messages in channels that have been
        designated as allowing pins. It checks if the channel is in the pin_cache
        and if the message isn't already pinned before attempting to pin it.

        Args:
            interaction: The interaction that triggered the command
            message: The message to pin
        """
        if message.channel.id in [x["id"] for x in self.pin_cache]:
            if not message.pinned:
                try:
                    await message.pin()
                except discord.Forbidden:
                    await interaction.response.send_message(
                        "I don't have permission to pin messages in this channel",
                        ephemeral=True,
                    )
                    return
                except discord.NotFound:
                    await interaction.response.send_message(
                        "I could not find that message", ephemeral=True
                    )
                    return
                except discord.HTTPException:
                    await interaction.response.send_message(
                        "Failed to pin the message. There are probably too many pins in this channel",
                        ephemeral=True,
                    )
                    return
                await interaction.response.send_message(f"Message pinned")
            else:
                await interaction.response.send_message(
                    "That message is already pinned", ephemeral=True
                )
        else:
            await interaction.response.send_message(
                "You can't pin messages in this channel", ephemeral=True
            )

    # Set which channels the pin command should work in
    @app_commands.checks.has_permissions(ban_members=True)
    @app_commands.default_permissions(ban_members=True)
    @app_commands.command(
        name="set_pin_channels",
        description="Set which channels the pin command should work in",
    )
    async def set_pin_channels(
        self, interaction: discord.Interaction, channel: discord.TextChannel
    ) -> None:
        """
        Toggle whether the pin command can be used in a specific channel.

        This admin-only command toggles whether users can use the pin context menu
        command in a specific channel. If the channel is already in the allowed list,
        it will be removed; otherwise, it will be added.

        Args:
            interaction: The interaction that triggered the command
            channel: The channel to toggle pin permissions for
        """
        if channel.id in [x["id"] for x in self.pin_cache]:
            await self.bot.db.execute(
                "UPDATE channels SET allow_pins = FALSE WHERE id = $1", channel.id
            )
            await interaction.response.send_message(
                f"Removed {channel.name} to allowed pin channels", ephemeral=True
            )
            self.pin_cache = await self.bot.db.fetch(
                "SELECT id FROM channels where allow_pins = TRUE"
            )
        else:
            await self.bot.db.execute(
                "UPDATE channels SET allow_pins = TRUE WHERE id = $1", channel.id
            )
            await interaction.response.send_message(
                f"Added {channel.name} to allowed pin channels", ephemeral=True
            )
            self.pin_cache = await self.bot.db.fetch(
                "SELECT id FROM channels where allow_pins = TRUE"
            )


async def setup(bot: commands.Bot) -> None:
    """
    Set up the OtherCogs cog.

    This function is called automatically by the bot when loading the extension.

    Args:
        bot: The bot instance to attach the cog to
    """
    await bot.add_cog(OtherCogs(bot))
